from flask import Flask, render_template, request, jsonify, send_from_directory
from flask_cors import CORS
from datetime import datetime, date
import os
from openpyxl import load_workbook
from config import Config
from models import db, AdData, RevenueData

# 初始化Flask应用
app = Flask(__name__)
app.config.from_object(Config)
CORS(app)

# 初始化数据库
db.init_app(app)

# 创建上传目录
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# ==================== 辅助函数 ====================

def parse_date(date_str):
    """解析日期字符串"""
    if isinstance(date_str, date):
        return date_str
    return datetime.strptime(date_str, '%Y-%m-%d').date()

def build_ad_query(filters):
    """构建投放数据查询"""
    query = AdData.query
    
    if filters.get('startDate'):
        query = query.filter(AdData.date >= parse_date(filters['startDate']))
    if filters.get('endDate'):
        query = query.filter(AdData.date <= parse_date(filters['endDate']))
    if filters.get('pkg'):
        query = query.filter(AdData.pkg == filters['pkg'])
    if filters.get('group'):
        query = query.filter(AdData.group_name == filters['group'])
    if filters.get('panel'):
        query = query.filter(AdData.panel == filters['panel'])
    
    return query

# ==================== 前端页面路由 ====================

@app.route('/')
def index():
    """首页"""
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    """数据看板页面"""
    return render_template('dashboard.html')

@app.route('/history')
def history():
    """历史记录页面"""
    return render_template('history.html')

# ==================== API路由 ====================

@app.route('/api/health')
def health_check():
    """健康检查"""
    return jsonify({'status': 'ok', 'timestamp': datetime.now().isoformat()})

# ==================== 投放数据 API ====================

@app.route('/api/ads', methods=['GET'])
def get_ads():
    """获取投放数据"""
    try:
        filters = {
            'startDate': request.args.get('startDate'),
            'endDate': request.args.get('endDate'),
            'pkg': request.args.get('pkg'),
            'group': request.args.get('group'),
            'panel': request.args.get('panel')
        }
        
        query = build_ad_query(filters)
        ads = query.order_by(AdData.date.desc(), AdData.id.desc()).all()
        
        return jsonify({'success': True, 'data': [ad.to_dict() for ad in ads]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/ads', methods=['POST'])
def create_ad():
    """创建投放数据"""
    try:
        data = request.json
        ad = AdData(
            date=parse_date(data['date']),
            pkg=data['pkg'],
            group_name=data['group'],
            panel=data['panel'],
            spend=float(data.get('spend', 0)),
            spend_clean=float(data.get('spendClean', 0)),
            revenue=float(data.get('revenue', 0))
        )
        db.session.add(ad)
        db.session.commit()
        
        return jsonify({'success': True, 'id': ad.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/ads/batch', methods=['POST'])
def create_ads_batch():
    """批量创建投放数据"""
    try:
        data_list = request.json.get('data', [])
        ads = []
        
        for item in data_list:
            ad = AdData(
                date=parse_date(item['date']),
                pkg=item['pkg'],
                group_name=item['group'],
                panel=item['panel'],
                spend=float(item.get('spend', 0)),
                spend_clean=float(item.get('spendClean', 0)),
                revenue=float(item.get('revenue', 0))
            )
            ads.append(ad)
        
        db.session.bulk_save_objects(ads)
        db.session.commit()
        
        return jsonify({'success': True, 'count': len(ads)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/ads/<int:ad_id>', methods=['DELETE'])
def delete_ad(ad_id):
    """删除投放数据"""
    try:
        ad = AdData.query.get_or_404(ad_id)
        db.session.delete(ad)
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/ads/delete-batch', methods=['POST'])
def delete_ads_batch():
    """批量删除投放数据"""
    try:
        ids = request.json.get('ids', [])
        AdData.query.filter(AdData.id.in_(ids)).delete()
        db.session.commit()
        
        return jsonify({'success': True, 'deleted': len(ids)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== 收入数据 API ====================

@app.route('/api/revenues', methods=['GET'])
def get_revenues():
    """获取收入数据"""
    try:
        filters = {
            'startDate': request.args.get('startDate'),
            'endDate': request.args.get('endDate'),
            'installDays': request.args.get('installDays')
        }
        
        query = RevenueData.query
        
        if filters['startDate']:
            query = query.filter(RevenueData.date >= parse_date(filters['startDate']))
        if filters['endDate']:
            query = query.filter(RevenueData.date <= parse_date(filters['endDate']))
        if filters['installDays']:
            query = query.filter(RevenueData.install_days == filters['installDays'])
        
        revenues = query.order_by(RevenueData.date.desc(), RevenueData.id.desc()).all()
        
        return jsonify({'success': True, 'data': [rev.to_dict() for rev in revenues]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/revenues', methods=['POST'])
def create_revenue():
    """创建收入数据"""
    try:
        data = request.json
        revenue = RevenueData(
            date=parse_date(data['date']),
            install_days=data['installDays'],
            coin=float(data.get('coin', 0)),
            first_sub=float(data.get('firstSub', 0)),
            renew_sub=float(data.get('renewSub', 0)),
            coin_renew=float(data.get('coinRenew', 0)),
            total=float(data.get('total', 0))
        )
        db.session.add(revenue)
        db.session.commit()
        
        return jsonify({'success': True, 'id': revenue.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/revenues/batch', methods=['POST'])
def create_revenues_batch():
    """批量创建收入数据"""
    try:
        data_list = request.json.get('data', [])
        revenues = []
        
        for item in data_list:
            revenue = RevenueData(
                date=parse_date(item['date']),
                install_days=item['installDays'],
                coin=float(item.get('coin', 0)),
                first_sub=float(item.get('firstSub', 0)),
                renew_sub=float(item.get('renewSub', 0)),
                coin_renew=float(item.get('coinRenew', 0)),
                total=float(item.get('total', 0))
            )
            revenues.append(revenue)
        
        db.session.bulk_save_objects(revenues)
        db.session.commit()
        
        return jsonify({'success': True, 'count': len(revenues)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/revenues/<int:rev_id>', methods=['DELETE'])
def delete_revenue(rev_id):
    """删除收入数据"""
    try:
        revenue = RevenueData.query.get_or_404(rev_id)
        db.session.delete(revenue)
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/revenues/delete-batch', methods=['POST'])
def delete_revenues_batch():
    """批量删除收入数据"""
    try:
        ids = request.json.get('ids', [])
        RevenueData.query.filter(RevenueData.id.in_(ids)).delete()
        db.session.commit()
        
        return jsonify({'success': True, 'deleted': len(ids)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== 统计数据 API ====================

@app.route('/api/stats', methods=['GET'])
def get_stats():
    """获取统计数据"""
    try:
        filters = {
            'startDate': request.args.get('startDate'),
            'endDate': request.args.get('endDate'),
            'pkg': request.args.get('pkg'),
            'group': request.args.get('group'),
            'panel': request.args.get('panel')
        }
        
        query = build_ad_query(filters)
        ads = query.all()
        
        total_spend = sum(ad.spend for ad in ads)
        total_spend_clean = sum(ad.spend_clean for ad in ads)
        total_revenue = sum(ad.revenue for ad in ads)
        
        roi = (total_revenue / total_spend) if total_spend > 0 else 0
        roi_clean = (total_revenue / total_spend_clean) if total_spend_clean > 0 else 0
        
        return jsonify({
            'success': True,
            'data': {
                'totalSpend': total_spend,
                'totalSpendClean': total_spend_clean,
                'totalRevenue': total_revenue,
                'roi': roi,
                'roiClean': roi_clean
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/chart-data', methods=['GET'])
def get_chart_data():
    """获取图表数据"""
    try:
        filters = {
            'startDate': request.args.get('startDate'),
            'endDate': request.args.get('endDate'),
            'pkg': request.args.get('pkg'),
            'group': request.args.get('group'),
            'panel': request.args.get('panel')
        }
        
        query = build_ad_query(filters)
        ads = query.order_by(AdData.date.asc()).all()
        
        # 按日期汇总
        date_summary = {}
        for ad in ads:
            date_str = ad.date.strftime('%Y-%m-%d')
            if date_str not in date_summary:
                date_summary[date_str] = {'spend': 0, 'spendClean': 0, 'revenue': 0}
            date_summary[date_str]['spend'] += ad.spend
            date_summary[date_str]['spendClean'] += ad.spend_clean
            date_summary[date_str]['revenue'] += ad.revenue
        
        chart_data = []
        for date_str, summary in sorted(date_summary.items()):
            roi = (summary['revenue'] / summary['spend']) if summary['spend'] > 0 else 0
            roi_clean = (summary['revenue'] / summary['spendClean']) if summary['spendClean'] > 0 else 0
            chart_data.append({
                'date': date_str,
                'spend': summary['spend'],
                'spendClean': summary['spendClean'],
                'revenue': summary['revenue'],
                'roi': roi,
                'roiClean': roi_clean
            })
        
        return jsonify({'success': True, 'data': chart_data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== Excel导入API ====================

def read_excel_data(file):
    """读取Excel或CSV文件数据"""
    filename = file.filename
    
    if filename.endswith('.csv'):
        # 读取CSV
        import csv
        from io import StringIO
        stream = StringIO(file.read().decode('utf-8'))
        reader = csv.DictReader(stream)
        return list(reader)
    else:
        # 读取Excel
        from openpyxl import load_workbook
        wb = load_workbook(filename=file, data_only=True)
        ws = wb.active
        
        # 读取表头
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # 读取数据
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = value
            if any(row_dict.values()):  # 跳过空行
                data.append(row_dict)
        
        return data

@app.route('/api/import/ad', methods=['POST'])
def import_ad_excel():
    """导入投放数据Excel"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '没有选择文件'}), 400
        
        # 读取Excel数据
        data_list = read_excel_data(file)
        
        # 转换为数据列表
        ads = []
        for row in data_list:
            ad = AdData(
                date=parse_date(row.get('投放日期') or row.get('date')),
                pkg=str(row.get('包名') or row.get('pkg')),
                group_name=str(row.get('投放组') or row.get('group')),
                panel=str(row.get('面板') or row.get('panel')),
                spend=float(row.get('消耗') or row.get('spend') or 0),
                spend_clean=float(row.get('消耗去空耗') or row.get('spendClean') or 0),
                revenue=float(row.get('收入投放口径') or row.get('revenue') or 0)
            )
            ads.append(ad)
        
        # 批量插入
        db.session.bulk_save_objects(ads)
        db.session.commit()
        
        return jsonify({'success': True, 'count': len(ads)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/import/revenue', methods=['POST'])
def import_revenue_excel():
    """导入收入数据Excel"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '没有文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '没有选择文件'}), 400
        
        # 读取Excel数据
        data_list = read_excel_data(file)
        
        # 转换为数据列表
        revenues = []
        for row in data_list:
            revenue = RevenueData(
                date=parse_date(row.get('投放日期') or row.get('date')),
                install_days=str(row.get('天数') or row.get('installDays') or '当日'),
                coin=float(row.get('金币收入') or row.get('coin') or 0),
                first_sub=float(row.get('会员首订') or row.get('firstSub') or 0),
                renew_sub=float(row.get('会员续订') or row.get('renewSub') or 0),
                coin_renew=float(row.get('金币续订') or row.get('coinRenew') or 0)
            )
            revenue.total = revenue.coin + revenue.first_sub + revenue.renew_sub + revenue.coin_renew
            revenues.append(revenue)
        
        # 批量插入
        db.session.bulk_save_objects(revenues)
        db.session.commit()
        
        return jsonify({'success': True, 'count': len(revenues)})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== 宽表API ====================

@app.route('/api/wide-table', methods=['GET'])
def get_wide_table():
    """
    按日期聚合投放数据 + 各天数收入数据，返回大宽表
    列：日期 | 消耗 | 消耗(去空耗) | 收入(投放口径) | ROI | 去空耗ROI
        | 当日_金币收入 | 当日_会员首订 | 当日_会员续订 | 当日_收入合计
        | 8日_金币续订 | 8日_会员续订 | 8日_合计
        | 14日_金币续订 | 14日_会员续订 | 14日_合计
        | 30日_金币续订 | 30日_会员续订 | 30日_合计
        | 45日_金币续订 | 45日_会员续订 | 45日_合计
    """
    try:
        start_date = request.args.get('startDate')
        end_date = request.args.get('endDate')

        # 1. 按日期聚合投放数据
        ad_query = AdData.query
        if start_date:
            ad_query = ad_query.filter(AdData.date >= parse_date(start_date))
        if end_date:
            ad_query = ad_query.filter(AdData.date <= parse_date(end_date))
        ads = ad_query.all()

        ad_by_date = {}
        for ad in ads:
            d = ad.date.strftime('%Y-%m-%d')
            if d not in ad_by_date:
                ad_by_date[d] = {'spend': 0.0, 'spend_clean': 0.0, 'revenue': 0.0}
            ad_by_date[d]['spend'] += ad.spend
            ad_by_date[d]['spend_clean'] += ad.spend_clean
            ad_by_date[d]['revenue'] += ad.revenue

        # 2. 按日期+天数聚合收入数据
        rev_query = RevenueData.query
        if start_date:
            rev_query = rev_query.filter(RevenueData.date >= parse_date(start_date))
        if end_date:
            rev_query = rev_query.filter(RevenueData.date <= parse_date(end_date))
        revenues = rev_query.all()

        # rev_by_date_days[date][days] = {coin, firstSub, renewSub, coinRenew, total}
        rev_by_date = {}
        for rev in revenues:
            d = rev.date.strftime('%Y-%m-%d')
            days = str(rev.install_days)
            if d not in rev_by_date:
                rev_by_date[d] = {}
            if days not in rev_by_date[d]:
                rev_by_date[d][days] = {'coin': 0.0, 'firstSub': 0.0, 'renewSub': 0.0, 'coinRenew': 0.0, 'total': 0.0}
            rev_by_date[d][days]['coin'] += rev.coin
            rev_by_date[d][days]['firstSub'] += rev.first_sub
            rev_by_date[d][days]['renewSub'] += rev.renew_sub
            rev_by_date[d][days]['coinRenew'] += rev.coin_renew
            rev_by_date[d][days]['total'] += rev.total

        # 3. 合并所有日期
        all_dates = sorted(set(list(ad_by_date.keys()) + list(rev_by_date.keys())), reverse=True)

        # Use pure ASCII keys to avoid encoding issues on Windows
        # day0=当日, day8=8日, day14=14日, day30=30日, day45=45日
        DAYS_LIST = ['0', '8', '14', '30', '45']
        DAYS_KEY  = {'0': 'day0', '8': 'day8', '14': 'day14', '30': 'day30', '45': 'day45'}

        rows = []
        for d in all_dates:
            ad = ad_by_date.get(d, {'spend': 0.0, 'spend_clean': 0.0, 'revenue': 0.0})
            spend = ad['spend']
            spend_clean = ad['spend_clean']
            revenue = ad['revenue']
            roi = revenue / spend if spend > 0 else None
            roi_clean = revenue / spend_clean if spend_clean > 0 else None

            row = {
                'date': d,
                'spend': spend,
                'spendClean': spend_clean,
                'revenue': revenue,
                'roi': roi,
                'roiClean': roi_clean,
            }

            # Revenue fields per day segment
            for days in DAYS_LIST:
                key = DAYS_KEY[days]
                rev_days = rev_by_date.get(d, {}).get(days, {})
                if days == '0':
                    # Day-0: coin income + first sub + renew sub + total
                    row[f'{key}_coin'] = rev_days.get('coin', 0.0)
                    row[f'{key}_firstSub'] = rev_days.get('firstSub', 0.0)
                    row[f'{key}_renewSub'] = rev_days.get('renewSub', 0.0)
                    row[f'{key}_total'] = rev_days.get('total', 0.0)
                else:
                    # 8/14/30/45 days: coin renew + member renew + total
                    row[f'{key}_coinRenew'] = rev_days.get('coinRenew', 0.0)
                    row[f'{key}_renewSub'] = rev_days.get('renewSub', 0.0)
                    row[f'{key}_total'] = rev_days.get('total', 0.0)

            rows.append(row)

        return jsonify({'success': True, 'data': rows, 'total': len(rows)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 静态文件服务 ====================

@app.route('/static/<path:filename>')
def serve_static(filename):
    """服务静态文件"""
    return send_from_directory('static', filename)

# ==================== 初始化数据库 ====================

with app.app_context():
    db.create_all()
    print("数据库表已创建")

# ==================== 启动应用 ====================

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
